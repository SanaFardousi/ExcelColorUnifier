from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import zipfile

#The code from line 6 to 75 is from "https://gist.github.com/Mike-Honey/b36e651e9a7f1d2e1d60ce1c63b9b633"
from colorsys import rgb_to_hls, hls_to_rgb

#https://bitbucket.org/openpyxl/openpyxl/issues/987/add-utility-functions-for-colors-to-help
 
RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969
 
def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))
 
def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)
 
def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()
 
 
def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]
 
    colors = []
 
    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
 
        if 'window' in accent.getchildren()[0].attrib['val']:
            colors.append(accent.getchildren()[0].attrib['lastClr'])
        else:
            colors.append(accent.getchildren()[0].attrib['val'])
 
    return colors
 
def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))
 
def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))


# These are the codes for every shade of the coorisponding color in Excel in the theme : "Office 2013-2022"
# If your theme is different, this code won't work for you unless you find the color codes for your theme

blue = ['FF4472C4','FFB4C6E7','FF305496', 'FF5B9BD5', '44546A', '5B9BD5', '4470C4', 'D6DCE4', 'DFEBF7', 
        'DAE3F3', 'AEBACB', 'BDD7EE', 'B5C6E8', '8497B0', 
        '9DC3E6', '8EA9DB', '344050', '2E75B5', '2F5396', 
        '222A35', '1F4F7A', '203864', 'FF00B0F0', 'FF0070C0', 'FF002060', '5B9BD5', '1F497C', '4F80BD', '4AADC6', 'C7DAF1', 'DBE5F2', 'DBEFF4', '8DB4E3', 'B8CCE4', 'B6DEE8', '558ED5', '96B3D8', '93CEDD', '17375E', '376093', '31869B', '10253F', '254061', '215967']

green = ['70AD47', '548135', 'E2F0D9', '385724', 'C5E0B4', 'FF92D050', 'A9D18E', 'FF00B050', '9BBB59', 'C3D69B', 'EBF1DE', '77943C', 'D7E4BD', '4F6228']

orange = ['FFFFFF00','FFED7D31','ED7D32', 'FFBF00', 'FBE5D6', 'FFF2CC', 'F8CBAD', 'FFE699', 'F4B184', 'FFD966', 'C55911', 'BF8F00', '806000', 'F79546', 'FABF8F', 'FDEADA', 'E46C0A', 'FCD5B6', '994807']

red = ['FFC00000', 'FFFF0000', '853C0C', 'FFC0504D', 'D99694', 'F2DBDB', '953735', 'E6B9B8', '652523']

### Notice that some of the color codes have 6 characters and others have 8, this is because openpyxl treats Excel theme colors and standered colors differently
# read "https://stackoverflow.com/questions/72831896/openpyxl-workbook-theme-colors-not-standard-for-excel" 

# This function checks if the cell color matches any of the specified shades of red, blue, green, or orange.
# If a match is found, it changes the cell color to a specific shade of the identified color.

def unifyColors(cell, color):
    for r in red:
        if color == r:
            cell.fill = PatternFill(fill_type="solid", start_color="FF0000")
    for b in blue:
        if color == b:
            cell.fill = PatternFill(fill_type="solid", start_color="5B9BD5")
    for g in green:
        if color == g:
            cell.fill = PatternFill(fill_type="solid", start_color="92D050")
    for o in orange:
        if color == o:
            cell.fill = PatternFill(fill_type="solid", start_color="F4B084") 
    
# Extract the zip folder and modify each file 
zip = zipfile.ZipFile('sample files.zip')
nameList = zip.namelist()
zip.extractall()

for f in nameList:

    wb = load_workbook(f, data_only=True)
    ws = wb.active
    # Iterate through each cell and identify its color
    for row in ws.iter_rows(): 
        for cell in row:
            tint = cell.fill.start_color.tint
            index = cell.fill.start_color.index
            # Checks if the index is an integer, if it is, finds the color of the cell using the tint and index
            if(isinstance(index, int)):
                color = theme_and_tint_to_rgb(wb, int(index), tint)
                unifyColors(cell, color)
            else: # if it isn't an integer, then it's the color code 
                unifyColors(cell, index)

    wb.save(f)